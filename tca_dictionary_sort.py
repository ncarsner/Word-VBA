"""
sort_tca_dictionary.py

Sort a TCA dictionary workbook into the expected hierarchical order:

Title row first, then:
  Title, Chapter
  Title, Chapter, Part
  Title, Chapter, Part, Section

Sorting rules:
- Title, Chapter, Part, Section are treated as integers when present
- Rows are ordered by:
    (Title, Chapter, Part, Level, Section)
  where Level enforces:
    0 = Title-only
    1 = Chapter-only
    2 = Part-only
    3 = Section row

Notes:
- This script DOES NOT require pandas; it uses openpyxl only.
- It preserves the header row and rewrites the sheet contents.
- If multiple rows share the same sort key, original relative order is preserved (stable sort).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


@dataclass(frozen=True)
class RowSortKey:
    title: int
    chapter: int
    part: int
    level: int
    section: int


def _to_int_or_default(v: Any, default: int) -> int:
    if v is None:
        return default
    s = str(v).strip()
    if s == "":
        return default
    # handle accidental floats like "3.0"
    try:
        return int(float(s))
    except ValueError:
        return default


def _get_level(chapter: str, part: str, section: str) -> int:
    """
    0 = Title-only (no chapter/part/section)
    1 = Chapter-only
    2 = Part-only
    3 = Section
    """
    has_ch = chapter.strip() != ""
    has_pt = part.strip() != ""
    has_sc = section.strip() != ""

    if not has_ch and not has_pt and not has_sc:
        return 0
    if has_ch and not has_pt and not has_sc:
        return 1
    if has_ch and has_pt and not has_sc:
        return 2
    # Section rows (even if part is blank in the source, we still treat as section level)
    return 3


def sort_dictionary_sheet(
    xlsx_path: Path,
    sheet_name: str = "Dictionary",
    output_path: Path | None = None,
) -> Path:
    """
    Loads the workbook, sorts the rows on the dictionary sheet, and saves the result.

    If output_path is None, the input file is overwritten.
    """
    xlsx_path = xlsx_path.expanduser().resolve()
    if output_path is None:
        output_path = xlsx_path
    else:
        output_path = output_path.expanduser().resolve()

    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Read header
    header = [c.value for c in ws[1]]
    if not header or all(h is None for h in header):
        raise ValueError("Header row is missing or empty.")

    # Build column index map (case-insensitive)
    col_map: Dict[str, int] = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        col_map[str(h).strip().lower()] = idx

    required = ["jurisdiction", "title", "chapter", "part", "section", "value"]
    missing = [c for c in required if c not in col_map]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Found: {header}")

    # Extract rows (starting from row 2)
    data_rows: List[Tuple[RowSortKey, int, List[Any]]] = []
    # Include original index for stable sort
    original_index = 0

    for r in ws.iter_rows(min_row=2, values_only=False):
        values = [cell.value for cell in r]

        # Skip fully blank rows
        if all((v is None or str(v).strip() == "") for v in values):
            continue

        title_raw = values[col_map["title"]]
        chapter_raw = values[col_map["chapter"]]
        part_raw = values[col_map["part"]]
        section_raw = values[col_map["section"]]

        title_i = _to_int_or_default(title_raw, default=10**9)

        chapter_s = "" if chapter_raw is None else str(chapter_raw).strip()
        part_s = "" if part_raw is None else str(part_raw).strip()
        section_s = "" if section_raw is None else str(section_raw).strip()

        # Normalize Section: keep only digits if the sheet has hyphenated values by mistake
        # (does not change your source files; only affects sorting)
        if section_s and not section_s.isdigit():
            digits_only = "".join(ch for ch in section_s if ch.isdigit())
            section_s = digits_only

        level = _get_level(chapter_s, part_s, section_s)

        # For sorting: missing chapter/part/section should come *after* real values within same title
        chapter_i = _to_int_or_default(chapter_s, default=10**9)
        part_i = _to_int_or_default(part_s, default=10**9)
        section_i = _to_int_or_default(section_s, default=10**9)

        key = RowSortKey(
            title=title_i,
            chapter=chapter_i,
            part=part_i,
            level=level,
            section=section_i,
        )

        data_rows.append((key, original_index, values))
        original_index += 1

    # Sort (stable using original_index as final tie-breaker)
    data_rows.sort(key=lambda x: (x[0].title, x[0].chapter, x[0].part, x[0].level, x[0].section, x[1]))

    # Clear existing data (keep header row)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Write sorted rows back
    for _, __, row_values in data_rows:
        ws.append(row_values)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Example usage:
    #   python sort_tca_dictionary.py "TCA_title35_dictionary_only_hierarchical_order_FIXED.xlsx"
    import sys

    if len(sys.argv) < 2:
        print("Usage: python sort_tca_dictionary.py <path_to_xlsx> [sheet_name] [output_path]")
        raise SystemExit(2)

    in_path = Path(sys.argv[1])
    sheet = sys.argv[2] if len(sys.argv) >= 3 else "Dictionary"
    out_path = Path(sys.argv[3]) if len(sys.argv) >= 4 else None

    saved_to = sort_dictionary_sheet(in_path, sheet_name=sheet, output_path=out_path)
    print(f"Saved sorted workbook to: {saved_to}")
