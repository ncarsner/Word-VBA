"""
tca_dictionary_clean_validate_sort.py

One script that:
1) Validates a TCA “Dictionary” sheet for structural integrity
2) Sorts rows into the expected hierarchical order
3) Removes duplicates (exact duplicates by default)

Expected row hierarchy:
- Title row:         Title present, Chapter/Part/Section blank
- Chapter row:       Title + Chapter present, Part/Section blank
- Part row:          Title + Chapter + Part present, Section blank
- Section row:       Title + Chapter present, Section present (Part may be blank in some titles)

Sorting order:
  Title -> Chapter -> Part -> Level -> Section
  Level enforces: Title-only, Chapter-only, Part-only, Section

Duplicate removal:
- Default: exact duplicates across (Jurisdiction, Title, Chapter, Part, Section, Value, Status)
- Keeps the first occurrence and drops the rest
- Runs AFTER sorting (so the first occurrence is deterministic)

Dependencies:
- openpyxl (no pandas)

Usage:
  python tca_dictionary_clean_validate_sort.py "path/to/file.xlsx"
  python tca_dictionary_clean_validate_sort.py "path/to/file.xlsx" --sheet Dictionary --out "path/to/out.xlsx"
  python tca_dictionary_clean_validate_sort.py "path/to/file.xlsx" --inplace
  python tca_dictionary_clean_validate_sort.py "path/to/file.xlsx" --no-sort
  python tca_dictionary_clean_validate_sort.py "path/to/file.xlsx" --no-dedup
  python tca_dictionary_clean_validate_sort.py "path/to/file.xlsx" --strict-order

Exit codes:
  0 = success (may still have warnings)
  2 = validation errors found (file still written if --out/--inplace unless --fail-fast)

Notes:
- This script does not “fix” malformed values except:
  - It normalizes Section for sorting/validation purposes by stripping non-digits
  - It does NOT rewrite your Section values unless you enable --rewrite-section
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ----------------------------
# Helpers / normalization
# ----------------------------


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _to_int_or_default(v: Any, default: int) -> int:
    s = _s(v)
    if s == "":
        return default
    try:
        return int(float(s))
    except ValueError:
        return default


def _digits_only(s: str) -> str:
    return "".join(ch for ch in s if ch.isdigit())


def _norm_section_for_logic(section_raw: Any) -> str:
    """
    For validation/sorting logic:
    - Keep only digits
    - Accept 3-4 digits when present
    """
    s = _s(section_raw)
    if s == "":
        return ""
    if s.isdigit():
        return s
    return _digits_only(s)


def _level(chapter: str, part: str, section: str) -> int:
    """
    0 = Title-only
    1 = Chapter-only
    2 = Part-only
    3 = Section row
    """
    has_ch = chapter != ""
    has_pt = part != ""
    has_sc = section != ""
    if not has_ch and not has_pt and not has_sc:
        return 0
    if has_ch and not has_pt and not has_sc:
        return 1
    if has_ch and has_pt and not has_sc:
        return 2
    return 3


@dataclass(frozen=True)
class RowKey:
    jurisdiction: str
    title: str
    chapter: str
    part: str
    section: str
    value: str
    status: str


@dataclass(frozen=True)
class SortKey:
    title_i: int
    chapter_i: int
    part_i: int
    level: int
    section_i: int


# ----------------------------
# Core logic
# ----------------------------

REQUIRED_COLUMNS = ["jurisdiction", "title", "chapter", "part", "section", "value"]
OPTIONAL_COLUMNS = ["key", "status"]


def _build_colmap(header: List[Any]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        col_map[str(h).strip().lower()] = idx
    return col_map


def _read_sheet_rows(ws) -> Tuple[List[str], List[List[Any]]]:
    header = [c.value for c in ws[1]]
    if not header or all(h is None for h in header):
        raise ValueError("Header row is missing or empty.")
    header_norm = [("" if h is None else str(h).strip()) for h in header]

    rows: List[List[Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all((_s(v) == "") for v in r):
            continue
        rows.append(list(r))
    return header_norm, rows


def _validate_rows(
    rows: List[List[Any]],
    col: Dict[str, int],
    *,
    strict_order: bool = False,
) -> Tuple[List[str], List[str]]:
    """
    Returns (errors, warnings).

    strict_order:
      - if True, warn when the existing row order is not already hierarchical
        (useful to detect TOC contamination before sorting).
    """
    errors: List[str] = []
    warnings: List[str] = []

    # Basic per-row checks
    for i, r in enumerate(rows, start=2):
        jur = _s(r[col["jurisdiction"]]).upper()
        if jur and jur != "TCA":
            errors.append(f"Row {i}: Jurisdiction='{jur}' (expected 'TCA').")

        title = _s(r[col["title"]])
        if title == "":
            errors.append(f"Row {i}: Missing Title.")
        elif not _digits_only(title).isdigit():
            errors.append(f"Row {i}: Title='{title}' is not numeric.")

        ch = _s(r[col["chapter"]])
        pt = _s(r[col["part"]])

        sec_raw = r[col["section"]]
        sec = _norm_section_for_logic(sec_raw)

        # Section format (when present)
        if sec != "":
            if not sec.isdigit():
                errors.append(
                    f"Row {i}: Section='{_s(sec_raw)}' not numeric after normalization."
                )
            elif len(sec) not in (3, 4):
                warnings.append(f"Row {i}: Section='{sec}' is not 3–4 digits (check).")

        # Hierarchy sanity:
        # - If Part exists, Chapter must exist
        if pt != "" and ch == "":
            errors.append(f"Row {i}: Part='{pt}' present but Chapter is blank.")

        # - If Section exists, Chapter must exist
        if sec != "" and ch == "":
            errors.append(f"Row {i}: Section='{sec}' present but Chapter is blank.")

    if strict_order:
        # Validate the current ordering is already hierarchical:
        # based on numeric progression and level progression.
        last: Optional[Tuple[int, int, int, int, int]] = None

        for row_i, r in enumerate(rows, start=2):
            title_i = _to_int_or_default(r[col["title"]], 10**9)

            ch_s = _s(r[col["chapter"]])
            pt_s = _s(r[col["part"]])
            sec_s = _norm_section_for_logic(r[col["section"]])

            ch_i = _to_int_or_default(ch_s, 10**9)
            pt_i = _to_int_or_default(pt_s, 10**9)
            lvl = _level(ch_s, pt_s, sec_s)
            sec_i = _to_int_or_default(sec_s, 10**9)

            cur = (title_i, ch_i, pt_i, lvl, sec_i)
            if last is not None and cur < last:
                warnings.append(
                    f"Row {row_i}: Ordering appears non-hierarchical "
                    f"(row sort key {cur} < previous {last})."
                )
            last = cur

    return errors, warnings


def _make_sort_key(r: List[Any], col: Dict[str, int]) -> SortKey:
    title_i = _to_int_or_default(r[col["title"]], 10**9)

    ch_s = _s(r[col["chapter"]])
    pt_s = _s(r[col["part"]])
    sec_s = _norm_section_for_logic(r[col["section"]])

    ch_i = _to_int_or_default(ch_s, 10**9)
    pt_i = _to_int_or_default(pt_s, 10**9)
    lvl = _level(ch_s, pt_s, sec_s)
    sec_i = _to_int_or_default(sec_s, 10**9)

    return SortKey(
        title_i=title_i, chapter_i=ch_i, part_i=pt_i, level=lvl, section_i=sec_i
    )


def _dedup_rows(
    rows: List[List[Any]],
    col: Dict[str, int],
) -> Tuple[List[List[Any]], int]:
    """
    Removes exact duplicates using:
      (Jurisdiction, Title, Chapter, Part, Section, Value, Status)
    Keeps first.
    """
    seen = set()
    kept: List[List[Any]] = []
    removed = 0

    has_status = "status" in col

    for r in rows:
        key = RowKey(
            jurisdiction=_s(r[col["jurisdiction"]]).upper(),
            title=_s(r[col["title"]]),
            chapter=_s(r[col["chapter"]]),
            part=_s(r[col["part"]]),
            section=_norm_section_for_logic(r[col["section"]]),
            value=_s(r[col["value"]]),
            status=_s(r[col["status"]]) if has_status else "",
        )
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        kept.append(r)

    return kept, removed


def _rewrite_section_cells(rows: List[List[Any]], col: Dict[str, int]) -> int:
    """
    Optional: rewrite the sheet’s Section cell values to digits-only (3–4 digit final segment),
    based on the normalization used elsewhere.
    Returns count changed.
    """
    changed = 0
    s_idx = col["section"]
    for r in rows:
        original = _s(r[s_idx])
        normalized = _norm_section_for_logic(r[s_idx])
        if original != normalized:
            r[s_idx] = normalized
            changed += 1
    return changed


def process_workbook(
    in_path: Path,
    *,
    sheet_name: str,
    out_path: Optional[Path],
    inplace: bool,
    do_sort: bool,
    do_dedup: bool,
    strict_order: bool,
    rewrite_section: bool,
    fail_fast: bool,
) -> int:
    in_path = in_path.expanduser().resolve()
    if out_path is None and not inplace:
        out_path = in_path.with_name(in_path.stem + "_CLEANED.xlsx")
    if out_path is not None:
        out_path = out_path.expanduser().resolve()

    wb = load_workbook(in_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    header, rows = _read_sheet_rows(ws)
    col = _build_colmap(header)

    missing = [c for c in REQUIRED_COLUMNS if c not in col]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Found header: {header}")

    # Initial validation
    errors, warnings = _validate_rows(rows, col, strict_order=strict_order)

    if fail_fast and errors:
        for e in errors:
            print("ERROR:", e)
        for w in warnings:
            print("WARN:", w)
        return 2

    # Sort
    if do_sort:
        decorated = [(_make_sort_key(r, col), i, r) for i, r in enumerate(rows)]
        decorated.sort(
            key=lambda x: (
                x[0].title_i,
                x[0].chapter_i,
                x[0].part_i,
                x[0].level,
                x[0].section_i,
                x[1],
            )
        )
        rows = [r for _, __, r in decorated]

    # Deduplicate
    removed_dups = 0
    if do_dedup:
        rows, removed_dups = _dedup_rows(rows, col)

    # Optional rewrite of Section values
    section_changes = 0
    if rewrite_section:
        section_changes = _rewrite_section_cells(rows, col)

    # Re-validate after transformations (helpful to surface remaining issues)
    errors2, warnings2 = _validate_rows(rows, col, strict_order=False)
    errors.extend(errors2)
    warnings.extend(warnings2)

    # Rewrite sheet content (keep header, wipe existing data)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for r in rows:
        ws.append(r)

    # Save
    save_path = in_path if inplace else out_path
    assert save_path is not None
    wb.save(save_path)

    # Report
    print(f"Input:  {in_path}")
    print(f"Output: {save_path}")
    print(f"Rows written: {len(rows)}")
    if do_dedup:
        print(f"Duplicates removed: {removed_dups}")
    if rewrite_section:
        print(f"Section values rewritten (digits-only): {section_changes}")

    # Print top warnings/errors (no flooding)
    max_lines = 50
    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings[:max_lines]:
            print("WARN:", w)
        if len(warnings) > max_lines:
            print(f"WARN: ... {len(warnings) - max_lines} more")

    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors[:max_lines]:
            print("ERROR:", e)
        if len(errors) > max_lines:
            print(f"ERROR: ... {len(errors) - max_lines} more")
        return 2

    return 0


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", help="Path to the workbook")
    p.add_argument(
        "--sheet", default="Dictionary", help="Sheet name (default: Dictionary)"
    )
    p.add_argument(
        "--out",
        default=None,
        help="Output workbook path (default: <input>_CLEANED.xlsx)",
    )
    p.add_argument(
        "--inplace",
        action="store_true",
        help="Overwrite the input file (ignores --out)",
    )

    p.add_argument("--no-sort", action="store_true", help="Do not sort")
    p.add_argument("--no-dedup", action="store_true", help="Do not remove duplicates")
    p.add_argument(
        "--strict-order",
        action="store_true",
        help="Warn if input is not already hierarchical (pre-sort)",
    )

    p.add_argument(
        "--rewrite-section",
        action="store_true",
        help="Rewrite Section cells to digits-only normalization (optional)",
    )

    p.add_argument(
        "--fail-fast",
        action="store_true",
        help="If validation errors exist before changes, exit without writing output",
    )

    args = p.parse_args(argv)

    out_path = None if args.out is None else Path(args.out)

    return process_workbook(
        Path(args.xlsx),
        sheet_name=args.sheet,
        out_path=out_path,
        inplace=args.inplace,
        do_sort=not args.no_sort,
        do_dedup=not args.no_dedup,
        strict_order=args.strict_order,
        rewrite_section=args.rewrite_section,
        fail_fast=args.fail_fast,
    )


if __name__ == "__main__":
    raise SystemExit(main())
