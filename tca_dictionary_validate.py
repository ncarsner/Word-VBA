from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import openpyxl


EXPECTED_HEADERS = [
    "Jurisdiction",
    "Title",
    "Chapter",
    "Part",
    "Section",
    "KEY",
    "Value",
    "Status",
]

SECTION_RE = re.compile(r"^\d{3,4}$")  # only 3–4 digits, no hyphens


@dataclass
class Issue:
    file: str
    row: int
    column: str
    message: str

    def __str__(self) -> str:
        return f"{self.file} | row {self.row} | {self.column}: {self.message}"


def _norm_header(v: object) -> str:
    return str(v).strip() if v is not None else ""


def _to_int(v: object) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    s = str(v).strip()
    if s == "":
        return None
    if re.fullmatch(r"\d+", s):
        return int(s)
    return None


def validate_workbook(path: str) -> List[Issue]:
    issues: List[Issue] = []

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Dictionary" not in wb.sheetnames:
        issues.append(Issue(path, 0, "Sheet", "Missing required sheet: Dictionary"))
        return issues

    ws = wb["Dictionary"]

    # Build header index map from row 1
    header_row = 1
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        h = _norm_header(ws.cell(header_row, col).value)
        if h:
            header_map[h] = col

    # Minimum required headers
    for h in EXPECTED_HEADERS:
        if h not in header_map:
            issues.append(Issue(path, 1, "Header", f"Missing header: {h}"))

    if any(i.row == 1 and i.column == "Header" for i in issues):
        # Cannot reliably validate without headers
        return issues

    c_jur = header_map["Jurisdiction"]
    c_title = header_map["Title"]
    c_chap = header_map["Chapter"]
    c_part = header_map["Part"]
    c_sec = header_map["Section"]
    c_val = header_map["Value"]

    # Track hierarchy state while walking rows in order
    seen_title_row = False
    current_title: Optional[int] = None
    current_chapter: Optional[int] = None
    current_part: Optional[int] = None

    for r in range(2, ws.max_row + 1):
        jur = ws.cell(r, c_jur).value
        title = _to_int(ws.cell(r, c_title).value)
        chap = _to_int(ws.cell(r, c_chap).value)
        part = _to_int(ws.cell(r, c_part).value)
        sec_raw = ws.cell(r, c_sec).value
        val = ws.cell(r, c_val).value

        # skip completely blank rows
        if all(x is None or str(x).strip() == "" for x in [jur, title, chap, part, sec_raw, val]):
            continue

        # Jurisdiction must be exactly TCA
        if str(jur).strip() != "TCA":
            issues.append(Issue(path, r, "Jurisdiction", f"Expected 'TCA', got {jur!r}"))

        # Title required on every populated row (your files do this)
        if title is None:
            issues.append(Issue(path, r, "Title", "Title is required on populated rows"))
        else:
            # Detect and enforce "Title row first" principle
            if chap is None and part is None and sec_raw in (None, "", " "):
                # This is the Title row
                if seen_title_row and title != current_title:
                    # If you ever validate multi-title workbooks later, relax this.
                    issues.append(Issue(path, r, "Title", "Multiple Title rows detected; expected single-title workbook"))
                seen_title_row = True
                current_title = title
                current_chapter = None
                current_part = None
            else:
                if not seen_title_row:
                    issues.append(Issue(path, r, "Hierarchy", "Encountered Chapter/Part/Section before Title row"))

        # Section: must be blank or a 3–4 digit code (no hyphenation)
        sec: Optional[int] = None
        if sec_raw is not None and str(sec_raw).strip() != "":
            s = str(sec_raw).strip()
            if not SECTION_RE.fullmatch(s):
                issues.append(Issue(path, r, "Section", f"Section must be 3–4 digits only, got {sec_raw!r}"))
            else:
                sec = int(s)

        # Hierarchy rules
        # Chapter row: chap exists, part/section blank
        is_chapter_row = chap is not None and part is None and sec is None
        is_part_row = chap is not None and part is not None and sec is None
        is_section_row = chap is not None and sec is not None  # part may be None in some titles

        if is_chapter_row:
            # Chapter must not appear before title row
            if not seen_title_row:
                issues.append(Issue(path, r, "Hierarchy", "Chapter row appears before Title row"))
            # Update current state
            current_chapter = chap
            current_part = None

        if is_part_row:
            if current_chapter is None:
                issues.append(Issue(path, r, "Hierarchy", "Part row appears before any Chapter row"))
            elif chap != current_chapter:
                issues.append(Issue(path, r, "Hierarchy", f"Part row Chapter={chap} does not match current Chapter={current_chapter}"))
            current_part = part

        if is_section_row:
            if current_chapter is None:
                issues.append(Issue(path, r, "Hierarchy", "Section row appears before any Chapter row"))
            elif chap != current_chapter:
                issues.append(Issue(path, r, "Hierarchy", f"Section row Chapter={chap} does not match current Chapter={current_chapter}"))

            # If parts are in use for this chapter, require a part before sections
            if current_part is None:
                # Allow “no-part titles” (e.g., title 14/15 style) by only enforcing when the sheet contains any part rows.
                # We can detect that cheaply by checking if *any* part exists in the workbook.
                pass

            # If a part is present on the row, it must match current_part (after the part row)
            if part is not None:
                if current_part is None:
                    issues.append(Issue(path, r, "Hierarchy", "Section row has a Part value but no Part row has been set"))
                elif part != current_part:
                    issues.append(Issue(path, r, "Hierarchy", f"Section row Part={part} does not match current Part={current_part}"))

        # If part exists but chapter does not => invalid
        if chap is None and part is not None:
            issues.append(Issue(path, r, "Part", "Part provided but Chapter is blank"))

        # If section exists but chapter does not => invalid
        if chap is None and sec is not None:
            issues.append(Issue(path, r, "Section", "Section provided but Chapter is blank"))

        # Value required on populated rows (you can relax if you want)
        if val is None or str(val).strip() == "":
            issues.append(Issue(path, r, "Value", "Value is blank on a populated row"))

    # Additional: if any Part rows exist, enforce that sections must be under a part once a part has started
    # (This flags chapters where you have parts but forgot the part row before sections.)
    has_any_parts = False
    for r in range(2, ws.max_row + 1):
        chap = _to_int(ws.cell(r, c_chap).value)
        part = _to_int(ws.cell(r, c_part).value)
        sec_raw = ws.cell(r, c_sec).value
        sec_present = sec_raw is not None and str(sec_raw).strip() != ""
        if chap is not None and part is not None and not sec_present:
            has_any_parts = True
            break

    if has_any_parts:
        # Walk again, enforcing “if a chapter uses parts, sections must follow a part row”
        current_chapter = None
        current_part = None
        chapter_has_parts: Dict[int, bool] = {}
        chapter_part_started: Dict[int, bool] = {}

        # detect which chapters have any part rows
        for r in range(2, ws.max_row + 1):
            chap = _to_int(ws.cell(r, c_chap).value)
            part = _to_int(ws.cell(r, c_part).value)
            sec_raw = ws.cell(r, c_sec).value
            sec_present = sec_raw is not None and str(sec_raw).strip() != ""
            if chap is not None and part is not None and not sec_present:
                chapter_has_parts[chap] = True

        for r in range(2, ws.max_row + 1):
            chap = _to_int(ws.cell(r, c_chap).value)
            part = _to_int(ws.cell(r, c_part).value)
            sec_raw = ws.cell(r, c_sec).value
            sec_present = sec_raw is not None and str(sec_raw).strip() != ""

            if chap is not None and part is None and not sec_present:
                current_chapter = chap
                current_part = None

            if chap is not None and part is not None and not sec_present:
                current_chapter = chap
                current_part = part
                chapter_part_started[chap] = True

            if chap is not None and sec_present:
                if chapter_has_parts.get(chap, False) and not chapter_part_started.get(chap, False):
                    issues.append(Issue(path, r, "Hierarchy", f"Chapter {chap} uses Parts, but a Section appears before any Part row"))

    return issues


def iter_xlsx_paths(target: str) -> List[str]:
    if os.path.isdir(target):
        out = []
        for name in os.listdir(target):
            if name.lower().endswith(".xlsx"):
                out.append(os.path.join(target, name))
        return sorted(out)
    return [target]


def main() -> int:
    ap = argparse.ArgumentParser(description="Validate TCA Dictionary workbooks.")
    ap.add_argument("target", help="Path to a .xlsx workbook OR a folder containing .xlsx files")
    args = ap.parse_args()

    paths = iter_xlsx_paths(args.target)
    if not paths:
        print("No .xlsx files found.")
        return 2

    all_issues: List[Issue] = []
    for p in paths:
        try:
            issues = validate_workbook(p)
            all_issues.extend(issues)
        except Exception as e:
            all_issues.append(Issue(p, 0, "Exception", str(e)))

    if not all_issues:
        print("OK: no issues found.")
        return 0

    print(f"FOUND {len(all_issues)} ISSUE(S):")
    for iss in all_issues:
        print(str(iss))

    return 1


if __name__ == "__main__":
    raise SystemExit(main())
